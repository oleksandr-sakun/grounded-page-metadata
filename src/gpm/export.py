"""Export. Column names and sheet names are contractual -- treat as frozen."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

METADATA_COLUMNS = [
    "Order", "Asset Page #", "Asset Filename Placeholder", "Printed Page #",
    "Page Title", "Names Referenced", "Name Review",
    "Greek Chapter Referenced", "Chapter Review", "Subjects",
    "Suggested Subjects", "Notes for Review",
]

SHEETS = [
    "Metadata", "Name_Match_Audit", "Name_Review_Detail", "Chapter_Audit",
    "Chapter_Review_Detail", "QA_Flags", "Run_Notes",
]

REVIEW_FILL = PatternFill("solid", fgColor="FFF3CD")


def _autoformat(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, 60)


def write_workbook(path: Path, frames: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for sheet in SHEETS:
            df = frames.get(sheet, pd.DataFrame())
            if df.empty:
                df = pd.DataFrame(columns=["(no rows)"])
            df.to_excel(xw, sheet_name=sheet, index=False)
            _autoformat(xw.sheets[sheet])

        # Highlight rows a human actually has to look at.
        ws = xw.sheets["Metadata"]
        meta = frames["Metadata"]
        for i, (_, r) in enumerate(meta.iterrows(), start=2):
            if str(r["Name Review"]) or str(r["Chapter Review"]):
                for c in ws[i]:
                    c.fill = REVIEW_FILL


def batch_summary(rows: list[dict], path: Path) -> pd.DataFrame:
    df = pd.DataFrame(rows, columns=[
        "Source PDF filename", "Detected page count", "Metadata rows exported",
        "Name Review count", "Chapter Review count", "Dense page flags",
        "Invalid name count", "Invalid chapter count", "Output workbook path",
        "QA status", "Notes",
    ])
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="Batch_Summary", index=False)
        _autoformat(xw.sheets["Batch_Summary"])
    return df
